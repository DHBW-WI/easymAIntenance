import http.client
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill


#API call and recieving of the data
conn = http.client.HTTPSConnection("admin-api.axissecurity.com")
payload = ''
headers = {
  'Accept': 'application/json',
  'Authorization': 'Bearer eyJhbGciOiJSUzI1NiIsImtpZCI6IjM1MDEyRjQxMzVENEE3MURCQjQ3QjQ4RTk2NkY0MDYyNzUxREEzMUQiLCJ0eXAiOiJhdCtqd3QiLCJ4NXQiOiJOUUV2UVRYVXB4MjdSN1NPbG05QVluVWRveDAifQ.eyJuYmYiOjE3MDc4MTE1NzAsImV4cCI6MTcyMzM2MzU2OSwiaXNzIjoiaHR0cDovL29hdXRoIiwiYXhpc1Rva2VuSWQiOiIzZGJmMGUwMi1lMTExLTRjN2ItYmM1Mi0xMjE3ODI3NDhkMTYiLCJheGlzVG9rZW5OYW1lIjoicmVtb3RlX21haW50ZW5hbmNlX2FwaSIsImF4aXNUZW5hbnROYW1lIjoiSFBFIEVkZ2VDb25uZWN0IC0gRU1FQSIsImF1ZCI6InB1YmxpYy1hZG1pbi1hcGkiLCJ0ZW5hbnQiOiJmZTg2ZjAxZC03YjBjLTQ0MGMtOGFlZS01Yzg0MTc0ZWMwZWUiLCJzY29wZSI6WyJ1c2VyLnJlYWQiLCJncm91cC5yZWFkIiwicG9saWN5LnJlYWQiLCJ1c2VyLndyaXRlIiwiZ3JvdXAud3JpdGUiLCJwb2xpY3kud3JpdGUiXX0.VtNqt8vlu-W8scw7YTjUa6KlIfHnZ1EqOimZmc8t1VBM6jjBUMARYBS1A7H3bntxvGEHND57gV742gthTxY9nKus7OWezHm8X9L9y4_l99UX0EYguB042TmqlKMnGXEAKn19vbMusvfsqQ79g4b3S4gIrxJRvMa9vxF25Xc11scYuMunIQBcG1SDeq3TosqeaCNi301th950Gkp_sm1axMpOdEJvlenaKZN8I9cxBgeruldKYjwLvy6yYfSmWT9MfgdZrGpaawzbJp60v67siVFyZomF7mko9BbxwyCwsPd6SMNoYsT8TiiGppP4RIPlgwChDU0yJLzUa_Y5aKVrhg', #Add your API Token here
  'Cookie': 'ax_session=1699630990.192.38.959631|4167ccfb2b9a7f3e6a491af5764a54a0'
}
conn.request("GET", "/api/v1.0/Users?pageNumber=1&pageSize=150", payload, headers)
res = conn.getresponse()
data = res.read().decode("utf-8")
json_data = json.loads(data)
final_data = json_data['data']

#flatten the json data
rows = []

for record in final_data:
    id = record['id']
    username = record['userName']
    email = record['email']
    firstName = record['firstName']
    lastName = record['lastName']
    groups = record['groups']

    rows.append([username, email, firstName, lastName, id, groups])

#write to xlsx
    #define workbook and do formatting
excel = Workbook()
table = excel.active
dark_grey = PatternFill(patternType='solid', fgColor='A6A6A6')
light_grey = PatternFill(patternType='solid', fgColor='D9D9D9')
table.column_dimensions['A'].width = 35
table.column_dimensions['B'].width = 35
table.column_dimensions['C'].width = 35
table.column_dimensions['D'].width = 35
table.column_dimensions['E'].width = 35
table.column_dimensions['F'].width = 100

    #write header
header = ["Username", "Email", "First Name", "Last Name", "ID", "Groups"]
for s in range(0,6):
        table.cell(row=1,column=s+1).value = header[s]
        table.cell(row=1,column=s+1).fill = dark_grey

    #write the rows
for r in range(1,len(rows)):
    for s in range(0,6):
        string = str(rows[r][s])
        table.cell(row=r+1,column=s+1).value = string
        if s%2 == 1:
              table.cell(row=r+1,column=s+1).fill = light_grey
    
    #save to xlsx
excel.save('axis_users.xlsx')